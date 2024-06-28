import openpyxl.worksheet.worksheet
import openpyxl
from openpyxl import load_workbook
from changeForm import change_form
def validationCheck(ws1:openpyxl.worksheet.worksheet.Worksheet, ws2:openpyxl.worksheet.worksheet.Worksheet) -> None:
    if ws1.max_row != ws2.max_row or ws1.max_column != ws2.max_column:
        print("Exception: Worksheets have different dimensions.")
        return

    sheet_len = ws1.max_row
    # sheet의 row 길이만큼 순회
    for row_idx in range(1, sheet_len + 1):
        row1 = list(ws1.rows)[row_idx - 1]
        row2 = list(ws2.rows)[row_idx - 1]

        for cell1, cell2 in zip(row1, row2):
            if cell1.value != cell2.value:
                print(f"Difference found at Sheet: {ws2.title}, Row {row_idx}, Column {cell2.column_letter}:")
                print('[change result]')
                print(cell1.value)
                print('[check_target]')
                print(cell2.value)

excel_path = '/Users/dataly/Desktop/1E_1S_요약_평가_수정.xlsx'
save_path = '/Users/dataly/Desktop/1E_1S_요약_평가.xlsx'

sheet_base = 'SUMVAL_'
sheet_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

# 변환한 워크북
change_wb = change_form(excel_path)
# 검사대상 워크북
target_wb = load_workbook(save_path)

# 출력
print('#'*50)
print('Validation Check start')
print('#'*50)
print()
for sheet_name in sheet_names:
    validationCheck(change_wb[sheet_base+sheet_name], target_wb[sheet_base+sheet_name])
    print(sheet_base + sheet_name + ' Validation Check Done.')
    print('='*50)
print()
print('#'*50)