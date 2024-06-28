import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Color

# 색상을 채우고 테두리를 그리는 함수
def style(ws:openpyxl.worksheet.worksheet.Worksheet, start_row: int, start_col: int, end_row: int, end_col: int, rmflag: int) -> None:
    # 색상 및 테두리 스타일 지정
    if rmflag == 0:
        fill = PatternFill(fill_type="solid", fgColor=Color('E2E2E2'))
    else:
        fill = PatternFill(fill_type="solid", fgColor=Color('FFFFFF'))
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 범위 내 셀에 색상 및 테두리 적용
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill
            cell.border = thin_border

def merge_cell(ws:openpyxl.worksheet.worksheet.Worksheet) -> None:
    start_row = None
    start_col = None
    end_row = None
    end_col = None

    rmcolor_row = None
    rmcolor_col = None

    # 색상채우기 및 테두리
    for row in ws.rows:
        for cell in row:
            if cell.value == '순번':
                start_row = cell.row
                start_col = cell.column

            elif cell.value == '총점':
                end_row = cell.row
                end_col = cell.column + 5

            elif cell.value == '내용 총점':
                rmcolor_row = cell.row
                rmcolor_col = cell.column
        if end_row is not None:
            style(ws, start_row, start_col, end_row, end_col, 0)
            style(ws, rmcolor_row, rmcolor_col, rmcolor_row, rmcolor_col + 5, 1)

            start_row = None
            start_col = None
            end_row = None
            end_col = None

    # merge
    idx = 1
    paragraph_size = 14
    sheet_len = ws.max_row
    while idx <= sheet_len:
        # 1부터 6회 반복
        for i in range(idx, idx+6):
            ws.merge_cells(f'A{i}:C{i}')

        # 1부터 3회 반복
        for i in range(idx, idx+3):
            ws.merge_cells(f'D{i}:G{i}')

        # 4부터 3회 반복
        for i in range(idx+3, idx+7):
            ws.merge_cells(f'D{i}:E{i}')
            ws.merge_cells(f'F{i}:G{i}')

        ws.merge_cells(f'A{idx+7}:A{idx+13}')

        merge_list = ['B', 'E', 'G']
        for a in merge_list:
            ws.merge_cells(f'{a}{idx+7}:{a}{idx+9}')

        ws.merge_cells(f'B{idx+10}:C{idx+10}')
        ws.merge_cells(f'B{idx+13}:C{idx+13}')

        idx += paragraph_size

def change_form(excel_path:str) -> openpyxl.Workbook:
    # 여러개의 sheet를 건너가며 데이터를 가져오기 위해 아래 문자열과 리스트 선언
    sheet_base = 'SUMVAL_'
    sheet_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

    # 기존의 excel 파일
    wb = load_workbook(excel_path, data_only=True)
    # 새 데이터를 저장할 workbook
    new_wb = Workbook()
    # 최초생성 워크시트 삭제
    new_wb.remove_sheet(new_wb['Sheet'])

    # 원본파일의 시트 개수만큼 반복
    for i in range(len(sheet_names)):
        ws = wb[sheet_base + sheet_names[i]]

        all_values = []
        # 시트 순회할 때 row 위치 추적을 위한 index
        row_idx = 1
        # sheet의 순번을 기록할 변수
        cnt = 1
        for row in ws.rows:
            row_value = []

            # 요약문 글자수 행을 추가할 list
            summary_len = []
            # 내용 총점 행을 추가할 list
            detail_total = []

            # 기존의 내용을 순회하며 데이터 생성
            for cell in row:
                row_value.append(cell.value)
                start_idx = row_idx - 2
                end_idx = row_idx
                if cell.value == '요약문':
                    summary_len = ['요약문 글자수', None, None,
                                   len(ws.cell(row=cell.row, column=cell.column + 3).value), None,
                                   len(ws.cell(row=cell.row, column=cell.column + 5).value), None, None]
                elif cell.value == '내용':
                    row_value[0] = cnt
                    cnt += 1
                elif cell.value == '논거 및 실천 방안':
                    detail_total = [None, '내용 총점', None, f'=SUM(D{start_idx}:D{end_idx})', None,
                                    f'=SUM(F{start_idx}:F{end_idx})', None, None]
                elif cell.value == '총점':
                    row_value = [None, '총점', None, f'=(SUM(D{start_idx-1}:D{end_idx-1})-5)*16.7/5', None, f'=(SUM(F{start_idx-1}:F{end_idx-1})-5)*16.7/5', None, None]
                    break
            row_idx += 1

            all_values.append(row_value)

            # 요약문 글자수와 내용 총점을 리스트에 추가
            if summary_len:
                all_values.append(summary_len)
                row_idx += 1
            elif detail_total:
                all_values.append(detail_total)
                row_idx += 1

        # 워크시트 생성
        new_ws = new_wb.create_sheet(title=sheet_base + sheet_names[i])

        # 생성한 데이터를 새 워크시트에 저장
        for row in all_values:
            new_ws.append(row)
        # 병합하고 스타일 지정
        merge_cell(new_ws)

    return new_wb



